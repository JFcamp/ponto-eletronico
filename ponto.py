import cv2
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime, timedelta
from PIL import Image as PILImage
import os
import tkinter as tk
from tkinter import simpledialog, messagebox, ttk

# Função para capturar a imagem do funcionário


def capture_image(employee_name):
    create_user_directory(employee_name)
    cap = cv2.VideoCapture(0)
    while True:
        ret, frame = cap.read()
        cv2.imshow('Pressione "s" para capturar a foto', frame)
        if cv2.waitKey(1) & 0xFF == ord('s'):
            image_path = f"./users/{employee_name}/{
                datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
            cv2.imwrite(image_path, frame)
            break
    cap.release()
    cv2.destroyAllWindows()
    return image_path

# Função para criar diretório do usuário


def create_user_directory(employee_name):
    user_directory = f"./users/{employee_name}"
    if not os.path.exists(user_directory):
        os.makedirs(user_directory)

# Função para redimensionar a imagem


def resize_image(image_path, width, height):
    img = PILImage.open(image_path)
    img = img.resize((width, height), PILImage.LANCZOS)
    resized_path = f"./users/{os.path.basename(image_path)}"
    img.save(resized_path)
    return resized_path

# Função para registrar o ponto no arquivo Excel


def register_time(employee_name, image_path, tipo_registro):
    file_name = "registro_ponto.xlsx"
    password = "cleria123"

    # Verifica se o arquivo já existe
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
    else:
        wb = Workbook()

    # Verifica se a aba do funcionário já existe
    if employee_name in wb.sheetnames:
        sheet = wb[employee_name]
    else:
        sheet = wb.create_sheet(employee_name)
        sheet.append(["Data/Hora", "Tipo", "Foto"])

    # Adiciona os registros de ponto
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # Redimensiona a imagem para 150x150 pixels
    resized_image_path = resize_image(image_path, 150, 150)
    img = XLImage(resized_image_path)
    sheet.append([current_time, tipo_registro, ""])

    # Coloca a imagem na célula correspondente
    row = sheet.max_row
    img.anchor = f'C{row}'
    sheet.add_image(img)

    # Ajusta a largura da coluna para caber a data/hora e tipo de registro
    sheet.column_dimensions[get_column_letter(1)].width = 20
    sheet.column_dimensions[get_column_letter(2)].width = 10
    # Ajusta a altura da linha para acomodar a imagem
    sheet.row_dimensions[row].height = 100

    # Protege a planilha com a senha
    for ws in wb.worksheets:
        ws.protection.set_password(password)
        ws.protection.sheet = True

    # Salva o arquivo Excel
    wb.save(file_name)

# Função para ler a última ação registrada


def read_last_action(employee_name):
    file_name = f"./users/{employee_name}_last_action.txt"
    if os.path.exists(file_name):
        with open(file_name, 'r') as file:
            return file.read().strip()
    return None

# Função para salvar a última ação registrada


def write_last_action(employee_name, action):
    file_name = f"./users/{employee_name}_last_action.txt"
    with open(file_name, 'w') as file:
        file.write(action)

# Função para calcular as horas trabalhadas


def calculate_worked_hours(employee_name):
    file_name = "registro_ponto.xlsx"
    if not os.path.exists(file_name):
        return None

    wb = openpyxl.load_workbook(file_name)
    if employee_name not in wb.sheetnames:
        return None

    sheet = wb[employee_name]
    entries = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        entries.append(row)

    worked_hours = {}
    daily_times = {}

    for entry in entries:
        datetime_entry = datetime.strptime(entry[0], "%Y-%m-%d %H:%M:%S")
        entry_type = entry[1]
        date_key = datetime_entry.strftime("%Y-%m-%d")

        if date_key not in daily_times:
            daily_times[date_key] = {"Entrada": None, "Saída": None}

        daily_times[date_key][entry_type] = datetime_entry

    for date_key, times in daily_times.items():
        if times["Entrada"] and times["Saída"]:
            worked_time = times["Saída"] - times["Entrada"]
            if date_key in worked_hours:
                worked_hours[date_key] += worked_time
            else:
                worked_hours[date_key] = worked_time

    return worked_hours

# Função para formatar timedelta em horas e minutos


def format_timedelta(td):
    total_seconds = int(td.total_seconds())
    hours, remainder = divmod(total_seconds, 3600)
    minutes, _ = divmod(remainder, 60)
    return f"{hours}h {minutes}m"

# Função para registrar horas trabalhadas no arquivo Excel


def register_worked_hours(employee_name, worked_hours):
    file_name = "registro_ponto.xlsx"
    password = "cleria123"

    wb = openpyxl.load_workbook(file_name)
    if "Horas Trabalhadas" not in wb.sheetnames:
        sheet = wb.create_sheet("Horas Trabalhadas")
        sheet.append(["Funcionário", "Data", "Horas Trabalhadas"])

    sheet = wb["Horas Trabalhadas"]

    # Registra horas trabalhadas
    for date, hours in worked_hours.items():
        sheet.append([employee_name, date, format_timedelta(hours)])

    # Protege a planilha com a senha
    for ws in wb.worksheets:
        ws.protection.set_password(password)
        ws.protection.sheet = True

    wb.save(file_name)

# Função principal


def main():
    # Cria a janela principal
    root = tk.Tk()
    root.withdraw()  # Esconde a janela principal

    # Tenta adicionar um ícone customizado, se disponível
    try:
        root.iconbitmap(default='icon.ico')
    except tk.TclError:
        print("Ícone personalizado não encontrado. Usando o ícone padrão.")

    # Define o estilo da interface
    style = ttk.Style(root)
    style.theme_use('clam')  # Utiliza o tema 'clam'

    # Centraliza as janelas flutuantes
    def center_window(window):
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry(f'{width}x{height}+{x}+{y}')

    # Solicita o nome do funcionário
    employee_name = simpledialog.askstring(
        "Registro de Ponto", "Digite seu nome:", parent=root)
    center_window(root)

    while True:
        last_action = read_last_action(employee_name)

        if last_action == "Entrada":
            tipo_registro = "Saída"
        else:
            tipo_registro = "Entrada"

        image_path = capture_image(employee_name)
        register_time(employee_name, image_path, tipo_registro)
        msgbox = tk.Toplevel(root)
        tk.Label(msgbox, text=f"Ponto de {
                 tipo_registro} registrado com sucesso!").pack(padx=20, pady=20)
        center_window(msgbox)
        msgbox.grab_set()
        msgbox.transient(root)

        # Salva a última ação registrada
        write_last_action(employee_name, tipo_registro)

        # Opções para o próximo registro
        response = messagebox.askyesno(
            "Registro de Ponto", "Deseja registrar o próximo ponto?", parent=root)

        if not response:
            break

    root.destroy()  # Fecha a janela principal ao sair

    # Calcula as horas trabalhadas
    worked_hours = calculate_worked_hours(employee_name)

    # Registra as horas trabalhadas no arquivo Excel
    register_worked_hours(employee_name, worked_hours)


if __name__ == "__main__":
    main()
